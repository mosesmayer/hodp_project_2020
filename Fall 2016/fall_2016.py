#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 11 00:03:41 2020

@author: alfiantjandra
"""
import pandas as pd
from openpyxl import load_workbook

workbook=load_workbook(filename="fall_2016.xlsx")

sheet = workbook.active

total_dict = {"General Education":0}
name_tracker = "General Education"
total_courses = 0

for value in sheet.iter_rows(min_row=2,max_row=10000,min_col=1,max_col=14,values_only=True):
    if value[0] == "Grand Total":
        break
    else: 
        total_courses += 1
        

for value in sheet.iter_rows(min_row=2,max_row=total_courses+1,min_col=1,max_col=14,values_only=True):
    if value[4] == name_tracker:
        total_dict[value[4]] += value[13]
    else:
        name_tracker = value[4]
        if value[4] in total_dict:
            total_dict[value[4]]+=value[13]
        else:
            total_dict[value[4]] = value[13]

""" Undergrad"""
total_undergrad = {"General Education":0}
name_tracker_2 = "General Education"
for value in sheet.iter_rows(min_row=2,max_row=total_courses+1,min_col=1,max_col=14,values_only=True):
    if value[4] == name_tracker_2:
        total_undergrad[value[4]] += value[6]
    else:
       name_tracker_2 = value[4]
       if value[4] in total_undergrad:
           total_undergrad[value[4]]+=value[6]
       else:
           total_undergrad[value[4]] = value[6]

""" Cross-Registration"""
total_cross = {"General Education":0}
name_tracker_3 = "General Education"
for value in sheet.iter_rows(min_row=2,max_row=total_courses+1,min_col=1,max_col=14,values_only=True):
    if value[4] == name_tracker_3:
        total_cross[value[4]] += value[9]
    else:
       name_tracker_3 = value[4]
       if value[4] in total_cross:
           total_cross[value[4]]+=value[9]
       else:
           total_cross[value[4]] = value[9]
           
           
print(pd.DataFrame(total_dict.items(), columns=['Department', 'Total Enrollment']))
print(pd.DataFrame(total_undergrad.items(), columns=['Department', 'Undergrad Enrollment']))
print(pd.DataFrame(total_cross.items(), columns=['Department', 'Cross-Registration']))