#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 21:05:52 2020

@author: alfiantjandra
"""

from openpyxl import load_workbook

workbook  =load_workbook(filename="sample.xlsx")

sheet = workbook.active

for value in sheet.iter_rows(min_row=2,max_row=3,min_col=1,max_col=3,values_only=True):
    print(value[0])
    