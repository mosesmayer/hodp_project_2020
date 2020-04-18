#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 23:52:06 2020

@author: alfiantjandra
"""
import pandas as pd
from openpyxl import load_workbook

workbook=load_workbook(filename="spring_2019.xlsx")

sheet = workbook.active



total_courses = 0

for value in sheet.iter_rows(min_row=5,max_row=10000,min_col=1,max_col=14,values_only=True):
    if value[0] == "Grand Total":
        break
    else: 
        total_courses += 1
        
""" Total Enrollment"""
total_dict = {"General Education":0}
name_tracker_1 = "General Education"
for value in sheet.iter_rows(min_row=5,max_row=total_courses+4,min_col=1,max_col=14,values_only=True):
    if value[4] == name_tracker_1:
        total_dict[value[4]] += value[13]
    else:
       name_tracker_1 = value[4]
       if value[4] in total_dict:
           total_dict[value[4]]+=value[13]
       else:
           total_dict[value[4]] = value[13]
               
""" Undergrad"""
total_undergrad = {"General Education":0}
name_tracker_2 = "General Education"
for value in sheet.iter_rows(min_row=5,max_row=total_courses+4,min_col=1,max_col=14,values_only=True):
    if value[4] == name_tracker_2:
        total_undergrad[value[4]] += value[6]
    else:
       name_tracker_2 = value[4]
       if value[4] in total_undergrad:
           total_undergrad[value[4]]+=value[6]
       else:
           total_undergrad[value[4]] = value[6]

""" Cross-Registration"""
total_cross = {"General Education":0}
name_tracker_3 = "General Education"
for value in sheet.iter_rows(min_row=5,max_row=total_courses+4,min_col=1,max_col=14,values_only=True):
    if value[4] == name_tracker_3:
        total_cross[value[4]] += value[9]
    else:
       name_tracker_3 = value[4]
       if value[4] in total_cross:
           total_cross[value[4]]+=value[9]
       else:
           total_cross[value[4]] = value[9]
           
print(pd.DataFrame(total_dict.items(), columns=['Department', 'Total Enrollment']))
print(pd.DataFrame(total_undergrad.items(), columns=['Department', 'Undergrad Enrollment']))
print(pd.DataFrame(total_cross.items(), columns=['Department', 'Cross-Registration']))
#tester = 0
#for i in total_cross:
#    tester += total_cross[i]
#print(tester)