#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 23:52:06 2020

@author: alfiantjandra
"""
import pandas as pd
from openpyxl import load_workbook
import sys
from collections import defaultdict
import matplotlib.pyplot as plt
import numpy as np

# Command line to run program: python3 Excel_to_Pandas.py "Spring 2016/spring_2016.xlsx" "Spring 2017/spring_2017.xlsx"
# "Spring 2018/spring_2018.xlsx""Spring 2019/spring_2019.xlsx" "Fall 2016/fall_2016.xlsx" "Fall 2017/fall_2017.xlsx"
# "Fall 2018/fall_2018.xlsx" "Fall 2019/fall_2019.xlsx"

import subprocess

# filenames = [
#     "Fall\\ 2016/fall_2016.xlsx",
#     "Fall\\ 2017/fall_2017.xlsx",
#     "Fall\\ 2018/fall_2018.xlsx",
#     "Fall\\ 2019/fall_2019.xlsx",
#     "Spring\\ 2016/spring_2016.xlsx",
#     "Spring\\ 2017/spring_2017.xlsx",
#     "Spring\\ 2018/spring_2018.xlsx",
#     "Spring\\ 2019/spring_2019.xlsx",
# ]
#
# for i in filenames:
#     cmd = "python3 Excel_to_Pandas.py {}".format(i)
#     print(cmd)
#     subprocess.call(cmd, shell=True)
#
# argc = len(sys.argv)
# argv = sys.argv
#
# if not (4 > argc > 1):
#     print("Usage: python3 [name_of_file]")
#     sys.exit()

fallCR = defaultdict(list)
springCR = defaultdict(list)
fallG = defaultdict(list)
springG = defaultdict(list)
fallU = defaultdict(list)
springU = defaultdict(list)
for arg in sys.argv[1:]:
    excel_filename = arg
    workbook = load_workbook(filename=excel_filename)
    sheet = workbook.active

    total_courses = 0
    min_courses_line = 2

    for value in sheet.iter_rows(min_row=1, max_row=10000, min_col=1, max_col=14, values_only=True):
        if value[0] == "Course ID":
            break
        else:
            min_courses_line += 1

    for value in sheet.iter_rows(min_row=5, max_row=10000, min_col=1, max_col=14, values_only=True):
        if value[0] == "Grand Total":
            break
        else:
            total_courses += 1

    """ Total Enrollment """
    total_dict = {"General Education": 0}
    name_tracker_1 = "General Education"
    for value in sheet.iter_rows(min_row=min_courses_line, max_row=total_courses + min_courses_line - 1, min_col=1,
                                 max_col=14, values_only=True):
        if value[4] == name_tracker_1:
            total_dict[value[4]] += value[13]
        else:
            name_tracker_1 = value[4]
            if value[4] in total_dict:
                total_dict[value[4]] += value[13]
            else:
                total_dict[value[4]] = value[13]

    """ Undergrad """
    total_undergrad = {"General Education": 0}
    name_tracker_2 = "General Education"
    for value in sheet.iter_rows(min_row=min_courses_line, max_row=total_courses + min_courses_line - 1, min_col=1,
                                 max_col=14, values_only=True):
        if value[4] == name_tracker_2:
            total_undergrad[value[4]] += value[6]
        else:
            name_tracker_2 = value[4]
            if value[4] in total_undergrad:
                total_undergrad[value[4]] += value[6]
            else:
                total_undergrad[value[4]] = value[6]

    """ Cross-Registration """
    total_cross = {"General Education": 0}
    name_tracker_3 = "General Education"
    for value in sheet.iter_rows(min_row=min_courses_line, max_row=total_courses + min_courses_line - 1, min_col=1,
                                 max_col=14, values_only=True):
        if value[4] == name_tracker_3:
            total_cross[value[4]] += value[9]
        else:
            name_tracker_3 = value[4]
            if value[4] in total_cross:
                total_cross[value[4]] += value[9]
            else:
                total_cross[value[4]] = value[9]

    """ Graduate students """
    total_grad = {"General Education": 0}
    name_tracker_4 = "General Education"
    for value in sheet.iter_rows(min_row=min_courses_line, max_row=total_courses + min_courses_line - 1, min_col=1,
                                 max_col=14, values_only=True):
        if value[4] == name_tracker_4:
            total_grad[value[4]] += value[7]
        else:
            name_tracker_4 = value[4]
            if value[4] in total_grad:
                total_grad[value[4]] += value[7]
            else:
                total_grad[value[4]] = value[7]

    total_frame = pd.DataFrame(total_dict.items(), columns=['Department', 'Total Enrollment'])
    undergrad_frame = pd.DataFrame(total_undergrad.items(), columns=['Department', 'Undergrad Enrollment'])
    crossreg_frame = pd.DataFrame(total_cross.items(), columns=['Department', 'Cross-Registration'])
    grad_frame = pd.DataFrame(total_grad.items(), columns=['Department', 'Grad Enrollment'])

    crossreg_sorted = crossreg_frame.sort_values(by="Cross-Registration", ascending=False)

    # print(total_frame)
    # print(undergrad_frame)
    # print(crossreg_frame)
    # print(crossreg_sorted)

    """ Top 5 """
    dep_names = total_frame['Department'].values

    # Crossreg
    top10_crossreg = crossreg_sorted.head(10)
    crrg_dep_titles = top10_crossreg["Department"].values.tolist()
    crrg_dep_values = top10_crossreg["Cross-Registration"].values.tolist()

    # Non-undergrad
    non_undergrad_series = (total_frame["Total Enrollment"]
                            - undergrad_frame["Undergrad Enrollment"])
    non_undergrad_count = non_undergrad_series.to_numpy()
    non_undergrad = pd.DataFrame({'Department': dep_names,
                                  'Non-undergrad': non_undergrad_count})
    top5_non_undergrad = non_undergrad.sort_values(by="Non-undergrad", ascending=False).head(10)

    # Graduate students
    top10_grad = grad_frame.sort_values(by="Grad Enrollment", ascending=False).head(10)
    grad_dep_titles = top10_grad["Department"].values.tolist()
    grad_dep_values = top10_grad["Grad Enrollment"].values.tolist()

    # Undergrad students
    top10_ugrad = undergrad_frame.sort_values(by="Undergrad Enrollment", ascending=False).head(10)
    ugrad_dep_titles = top10_ugrad["Department"].values.tolist()
    ugrad_dep_values = top10_ugrad["Undergrad Enrollment"].values.tolist()


    def writeOutput():
        """ File output below """
        if excel_filename[0] == "S":
            output_file_title = "top10txt/" + excel_filename[12:-5] + "_top10.txt"
        else:
            output_file_title = "top10txt/" + excel_filename[10:-5] + "_top10.txt"
        print(output_file_title)
        with open(output_file_title, 'w') as outfile:
            outfile.write("Crossreg top 10:\n")
            for i, (x, y) in enumerate(zip(crrg_dep_titles, crrg_dep_values)):
                outfile.write("{}: {:32s}  {}\n".format(i + 1, x, y))
            outfile.write('\n')
            outfile.write("Grad top 10:\n")
            for i, (x, y) in enumerate(zip(grad_dep_titles, grad_dep_values)):
                outfile.write("{}: {:32s}  {}\n".format(i + 1, x, y))
            outfile.write('\n')
            outfile.write("Ugrad top 10:\n")
            for i, (x, y) in enumerate(zip(ugrad_dep_titles, ugrad_dep_values)):
                outfile.write("{}: {:32s}  {}\n".format(i + 1, x, y))
            outfile.write('\n')


    def makedict():
        if excel_filename[0] == "S":
            for index, row in crossreg_frame.iterrows():
                springCR[row["Department"]].append(row["Cross-Registration"])
            for index, row in grad_frame.iterrows():
                springG[row["Department"]].append(row["Grad Enrollment"])
            for index, row in undergrad_frame.iterrows():
                springU[row["Department"]].append(row["Undergrad Enrollment"])
        else:
            for index, row in crossreg_frame.iterrows():
                fallCR[row["Department"]].append(row["Cross-Registration"])
            for index, row in grad_frame.iterrows():
                fallG[row["Department"]].append(row["Grad Enrollment"])
            for index, row in undergrad_frame.iterrows():
                fallU[row["Department"]].append(row["Undergrad Enrollment"])

    writeOutput()
    makedict()

springdataCR = sorted(springCR.items(), key=lambda dept: sum(dept[1])/4, reverse=True)
springdeptsCR = [springdataCR[i][0] for i in range(5)]
springnumsCR = [springdataCR[i][1] for i in range(5)]

springdataG = sorted(springG.items(), key=lambda dept: sum(dept[1])/4, reverse=True)
springdeptsG = [springdataG[i][0] for i in range(5)]
springnumsG = [springdataG[i][1] for i in range(5)]

springdataU = sorted(springU.items(), key=lambda dept: sum(dept[1])/4, reverse=True)
springdeptsU = [springdataU[i][0] for i in range(5)]
springnumsU = [springdataU[i][1] for i in range(5)]

falldataCR = sorted(fallCR.items(), key=lambda dept: sum(dept[1])/4, reverse=True)
falldeptsCR = [falldataCR[i][0] for i in range(5)]
fallnumsCR = [falldataCR[i][1] for i in range(5)]

falldataG = sorted(fallG.items(), key=lambda dept: sum(dept[1])/4, reverse=True)
falldeptsG = [falldataG[i][0] for i in range(5)]
fallnumsG = [falldataG[i][1] for i in range(5)]

falldataU = sorted(fallU.items(), key=lambda dept: sum(dept[1])/4, reverse=True)
falldeptsU = [falldataU[i][0] for i in range(5)]
fallnumsU = [falldataU[i][1] for i in range(5)]


def plotbar(dept, num, title):
    barwidth = 0.2

    # set heights of bars
    bars2016 = [num[i][0] for i in range(5)]
    bars2017 = [num[i][1] for i in range(5)]
    bars2018 = [num[i][2] for i in range(5)]
    bars2019 = [num[i][3] for i in range(5)]

    # Set positions of bars on X axis
    r0 = np.arange(len(bars2016))
    r1 = [x + barwidth for x in r0]
    r2 = [x + barwidth for x in r1]
    r3 = [x + barwidth for x in r2]

    plt.figure(figsize=(16*0.8, 9*0.8))

    plt.bar(r0, bars2016, color="#D2232A", width=barwidth, edgecolor="white", label="2016")
    plt.bar(r1, bars2017, color="#DA4B50", width=barwidth, edgecolor="white", label="2017")
    plt.bar(r2, bars2018, color="#AC1D23", width=barwidth, edgecolor="white", label="2018")
    plt.bar(r3, bars2019, color="#86171B", width=barwidth, edgecolor="white", label="2019")

    plt.xlabel("Department", fontweight="bold")
    plt.xticks([r + barwidth for r in np.arange(barwidth/2, len(bars2016))], dept)
    plt.title(title)

    plt.legend()
    plt.savefig("graphs-color/{}-bar.png".format(title))
    plt.clf()


def plotline(dept, num, title):
    x = [2016, 2017, 2018, 2019]
    colors = ["#D2232A", "#DA4B50", "#AC1D23", "#86171B", "#601014"]
    plt.figure(figsize=(16*0.8, 9*0.8))
    for i in range(len(num)):
        plt.plot(x, num[i], linewidth=2, label=dept[i], color=colors[i])

    plt.xlabel("Year", fontweight="bold")
    plt.ylabel("# of Students", fontweight="bold")
    plt.xticks(np.arange(2016, 2020, 1))
    plt.title(title)
    plt.legend()
    plt.savefig("graphs-color/{}-line.png".format(title))
    plt.clf()


plotbar(springdeptsCR, springnumsCR, "Spring Cross-Registration")
plotbar(springdeptsG, springnumsG, "Spring Graduate Students")
plotbar(springdeptsU, springnumsU, "Spring Undergraduate Students")
plotbar(falldeptsCR, fallnumsCR, "Fall Cross-Registration")
plotbar(falldeptsG, fallnumsG, "Fall Graduate Students")
plotbar(falldeptsU, fallnumsU, "Fall Undergraduate Students")

plotline(springdeptsCR, springnumsCR, "Spring Cross-Registration")
plotline(springdeptsG, springnumsG, "Spring Graduate Students")
plotline(springdeptsU, springnumsU, "Spring Undergraduate Students")
plotline(falldeptsCR, fallnumsCR, "Fall Cross-Registration")
plotline(falldeptsG, fallnumsG, "Fall Graduate Students")
plotline(falldeptsU, fallnumsU, "Fall Undergraduate Students")
